from openpyxl import load_workbook
import os

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')

file_excel_name = desktop_path + r'\F-1载荷位移.xlsx'
file_txt_name = desktop_path + r'\res-F-1载荷位移.csv'

wb = load_workbook(file_excel_name)
ws = wb.active

sheet_value = []
for cell in ws['Z']:
    if cell.value is not None:
        sheet_value.append(cell.value)

print(len(sheet_value))

with open(file=file_txt_name, mode='w', encoding='utf-8') as file_out:
    for i in range(0, len(sheet_value), 10):
        max_in_ten = sheet_value[i]
        for j in range(1, 10):
            if (i + j < len(sheet_value)):
                max_in_ten = max(max_in_ten, sheet_value[i + j])

        file_out.write(f'{max_in_ten}\n')
